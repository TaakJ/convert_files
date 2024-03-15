import glob, shutil
from pathlib import Path
import warnings
import logging
import pandas as pd
from verify_data import validate_files
from exception import CustomException
from datetime import datetime
from setup import Folder
from datetime import datetime
import openpyxl
from openpyxl.styles import Font
import shutil


class convert_2_file(validate_files):
    def __init__(self, **kwargs):
        super().__init__()

        self.__dict__.update(kwargs)

        for key, value in self.__dict__.items():
            setattr(self, key, value)

        self.__log = []
        self.date = datetime.now()

        self.get_list_files()
        self.get_data_files()
        self.write_data_to_tmp_file()
        self.write_data_to_target_file()

    @property
    def logging(self):
        return self.__log

    @logging.setter
    def logging(self, log):
        self.__log = log

    def check_success_files(func):
        def wrapper_success_files(*args):

            logging.info("Check Success files..")

            success_file = []
            for key in func(*args):
                filename = key['full_path']
                full_path = Folder.RAW + filename

                if glob.glob(full_path, recursive=True):
                    status = "successed"
                    success_file.append(status)
                else:
                    status = "missing"
                    success_file.append(status)
                key.update({'full_path': full_path, 'status': status})

            if success_file.__contains__("missing"):
                raise CustomException(errors=func(*args))
            else:
                logging.info(f"File found count {len(success_file)} status: successed.")

            return func(*args)
        return wrapper_success_files

    @check_success_files
    def get_list_files(self):
        if self.__log == []:
            args = []
            for file in Folder.FILE:
                source = Path(file).stem
                args.append({'source': source, 'full_path': file})
            self.__log = args

        return self.__log

    def mock_data(func):

        logging.info("Mock Data")

        def wrapper_mock_data(*args):
            mock_data = [['ApplicationCode',	'AccountOwner', 'AccountName',	'AccountType',	'EntitlementName',	'SecondEntitlementName','ThirdEntitlementName', 'AccountStatus',	'IsPrivileged',	'AccountDescription',
                        'CreateDate','LastLogin','LastUpdatedDate',	'AdditionalAttribute'],
                        [1,2,3,4,5,6,7,8,9,10,args[0].batch_date.strftime('%Y-%m-%d'),12, args[0].date,14],
                        # [15,16,17,18,19,20,21,22,23,24,args[0].batch_date.strftime('%Y-%m-%d'),26, args[0].date,28],
                        ]
            df = pd.DataFrame(mock_data)
            df.columns = df.iloc[0].values
            df = df[1:]
            df = df.reset_index(drop=True)
            func(*args).append({'source': 'Target_file', 'data': df.to_dict('list')})

        return wrapper_mock_data

    @mock_data
    def get_data_files(self):

        logging.info("Get Data from files..")

        status = "failed"
        for key in self.__log:
            full_path = key['full_path']
            types = Path(key['full_path']).suffix
            key.update({'status': status})

            try:
                if ['.xlsx', '.xls'].__contains__(types):
                    logging.info(f"Read Excel files: '{full_path}'.")
                    clean_data = self.generate_excel_data(full_path=full_path)
                else:
                    logging.info(f"Read Text files: '{full_path}'.")
                    clean_data = self.generate_text_data(full_path=full_path)

                status = "successed"
                key.update({'data': clean_data, 'status': status})

            except Exception as err:
                key.update({'errors': err})

            if "errors" in key:
                raise CustomException(errors=self.__log)

        return self.__log

    def write_data_to_tmp_file(self):

        logging.info("Write Data to Tmp files..")

        source_name = f"{Folder.TEMPLATE}Application Data Requirements.xlsx"
        tmp_name =  f"{Folder.TMP}TMP_{self.batch_date.strftime('%d%m%Y')}.xlsx"
        status = "failed"

        for key in self.__log:
            try:
                if key['source'] == "Target_file":
                    key.update({'full_path': tmp_name, 'status': status})
                    ## get new data.
                    new_df = pd.DataFrame(key["data"])
                    new_df['remark'] = "Inserted"
                    try:
                        workbook = openpyxl.load_workbook(tmp_name)
                        get_sheet = workbook.get_sheet_names()
                        sheet_num = len(get_sheet)
                        sheet_name = f"RUN_TIME_{sheet_num - 1}"
                        sheet = workbook.get_sheet_by_name(sheet_name)
                        workbook.active = sheet_num

                    except FileNotFoundError:
                        ## copy files from template.
                        status = self.copy_worksheet(source_name, tmp_name)
                        workbook = openpyxl.load_workbook(tmp_name)
                        sheet = workbook.worksheets[0]
                        sheet_name = "RUN_TIME_1"
                        sheet_num = 1
                        sheet.title = sheet_name

                    logging.info(f"Genarate Sheet_name: {sheet_name} in Tmp files.")

                    # read tmp files.
                    data = sheet.values
                    columns = next(data)[0:]
                    tmp_df = pd.DataFrame(data, columns=columns)

                    if status != "successed":
                        tmp_df = tmp_df.loc[tmp_df['remark'] != "Removed"]
                        ## create new shhet.
                        sheet_name = f"RUN_TIME_{sheet_num}"
                        sheet = workbook.create_sheet(sheet_name)
                    else:
                        tmp_df['remark'] = "Inserted"

                    new_data = self.validation_data(tmp_df, new_df)
                    ## write to tmp files.
                    status = self.write_worksheet(sheet, new_data)
                    workbook.move_sheet(workbook.active, offset=-sheet_num)
                    workbook.save(tmp_name)

                    key.update({'full_path': tmp_name, 'sheet_name': sheet_name,'status': status})
                    logging.info(f"Write to Tmp files status: {status}.")

            except Exception as err:
                key.update({'errors': err})

            if "errors" in key:
                raise CustomException(errors=self.__log)

    def write_worksheet(self, sheet, new_data):

        max_rows = max(new_data, default=0)
        logging.info(f"Data for write: {max_rows}. rows")
        start_rows = 2
        
        try:
            # write columns.
            for idx, columns in enumerate(new_data[start_rows].keys(), 1):
                sheet.cell(row=1, column=idx).value = columns
            ## write data.
            while start_rows <= max_rows:
                for remark in [new_data[start_rows][columns] for columns in new_data[start_rows].keys() if columns == 'remark']:
                    for idx, values in enumerate(new_data[start_rows].values(), 1):
                        if start_rows in self.skip_rows and remark == "Removed":
                            sheet.cell(row=start_rows, column=idx).value = values
                            sheet.cell(row=start_rows, column=idx).font = Font(bold=True, strike=True, color="00FF0000")
                            show = f"{remark} Rows: ({start_rows}) in Tmp files."
                        elif start_rows in self.upsert_rows.keys() and remark in ["Inserted", "Updated"]:
                            sheet.cell(row=start_rows, column=idx).value = values
                            show = f"{remark} Rows: ({start_rows}) in Tmp files. Record Changed: {self.upsert_rows[start_rows]}"
                        else:
                            sheet.cell(row=start_rows, column=idx).value = values
                            show = f"No Change Rows: ({start_rows}) in Tmp files."
                logging.info(show)
                start_rows += 1

            status = "successed"

        except KeyError as err:
            raise KeyError(f"Can not Wirte rows: {err} in Tmp files.")
        return status

    def copy_worksheet(self, source_name, target_name):
        try:
            if not glob.glob(target_name, recursive=True):
                shutil.copyfile(source_name, target_name)
            status = "successed"
        except FileNotFoundError as err:
            raise FileNotFoundError(err)

        return status

    def remove_row_empty(self, sheet):
        for row in sheet.iter_rows():
            if not all(cell.value for cell in row):
                sheet.delete_rows(row[0].row, 1)
                self.remove_row_empty(sheet)

    def write_data_to_target_file(self):

        logging.info("Write Data to Target files..")
        source_name = f"{Folder.TEMPLATE}Application Data Requirements.xlsx"
        target_name = f"{Folder.EXPORT}Application Data Requirements.xlsx"
        status = "failed"
        start_rows = 2

        for key in self.__log:
            try:
                if key['source'] == "Target_file":
                    ## read tmp file.
                    tmp_name = key['full_path']
                    sheet_name = key['sheet_name']
                    tmp_df = pd.read_excel(tmp_name, sheet_name=sheet_name)
                    tmp_df = tmp_df.loc[tmp_df['remark'] != "Removed"]

                    try:
                        ## check write mode.
                        if self.mode == "Append":
                            status = self.copy_worksheet(source_name, target_name)
                            if status == "successed":
                                workbook = openpyxl.load_workbook(target_name)
                                get_sheet = workbook.get_sheet_names()
                                sheet = workbook.get_sheet_by_name(get_sheet[0])
                                workbook.active = sheet
                        else:
                            self.mode = "Overwrite"
                            target_name = f"{Folder.EXPORT}{Path(target_name).stem}_{self.batch_date.strftime('%d%m%Y')}.xlsx"
                            status = "successed" if glob.glob(target_name, recursive=True) else self.copy_worksheet(source_name, target_name)

                            if status == "successed":
                                workbook = openpyxl.load_workbook(target_name)
                                get_sheet = workbook.get_sheet_names()
                                sheet = workbook.get_sheet_by_name(get_sheet[0])
                                workbook.active = sheet

                        ## read target file.
                        data = sheet.values
                        columns = next(data)[0:]
                        target_df = pd.DataFrame(data, columns=columns)
                        target_df['remark'] = "Inserted"

                        ## compare data tmp data with target data.
                        select_date = tmp_df['CreateDate'].unique()
                        status, new_data = self.customize_data(select_date, target_df, tmp_df)

                        key.update({'full_path': target_name, 'status': status})

                    except Exception as err:
                        raise Exception(err)

                    ## write data to target files.
                    logging.info(f"Write mode: {self.mode} in Terget_files: '{Path(target_name).name}'")
                    if status == "successed":
                        max_rows = max(new_data, default=0)
                        while start_rows <= max_rows:
                            for idx, columns in enumerate(new_data[start_rows].keys(), 1):
                                if columns == 'remark':
                                    if f'{start_rows}' in self.upsert_rows.keys() and new_data[start_rows][columns] in ["Updated", "Inserted"]:
                                        show = f"{new_data[start_rows][columns]} Rows: ({start_rows}) in Target files. Record Changed: {self.upsert_rows[f'{start_rows}']}"
                                    elif start_rows in self.skip_rows and new_data[start_rows][columns] == "Removed":
                                        show = f"{new_data[start_rows][columns]} Rows: ({start_rows}) in Target files."
                                        sheet.delete_rows(start_rows, sheet.max_row)
                                    else:
                                        show = f"No Change Rows: ({start_rows}) in Target files."
                                else:
                                    if start_rows in self.skip_rows:
                                        continue
                                    sheet.cell(row=start_rows, column=idx).value = new_data[start_rows][columns]
                                    continue
                                logging.info(show)
                            start_rows += 1
                    self.remove_row_empty(sheet)

                    ## save files.
                    workbook.save(target_name)
                    status = "successed"

                    key.update({'status': status})
                    logging.info(f"Write to Target Files status: {status}.")

            except Exception as err:
                key.update({'status': status, 'errors': err})

        if "errors" in key:
            raise CustomException(errors=self.__log)
